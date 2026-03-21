#!/usr/bin/env python3
"""
Monthly Budget Tracker
Reads RBC Visa (.csv) and Amex (.xls) statements from a folder,
categorizes expenses, and generates a monthly budget summary with
Excel output for Google Sheets.
"""

import csv
import json
import os
import re
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Missing dependency. Run:  pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    xlrd = None


# ── Configuration (edit these as your situation changes) ─────────

STIPEND = 3146.00
SAVINGS_RATE = 0.20
SAVINGS_GOAL = 20000.00
SAVINGS_BALANCE = 4400.00          # update this each month as you contribute
SAVINGS_DEADLINE = datetime(2027, 9, 1)

CATEGORIES = [
    "Work Travel (Island)",
    "Fixed Costs",
    "Food",
    "Other Necessary",
    "Discretionary/Fun",
]

# ── Auto-classification keywords (case-insensitive substring match) ──

AUTO_RULES = {
    "Work Travel (Island)": [
        "bc ferries", "bcf -", "bcf-",
        "chevron", "chv",
        "shell c0",
        "petro-canada", "petro canada",
        "squamish valley gas",
    ],
    "Fixed Costs": [
        "rent", "utilit", "wifi", "internet",
        "hydro", "bc hydro",
        "telus", "shaw", "rogers", "fido", "freedom mobile",
        "insurance", "bcaa",
        "classpass",
        "interest",
    ],
    "Food": [
        # Grocery stores
        "no frills", "superstore", "save-on", "safeway",
        "iga ", "walmart", "costco",
        "stong", "choices", "sungiven", "persia foods",
        "nesters", "mostafa", "grocery", "supermarket",
        "co-op",
        # Restaurants & cafes
        "cafe", "coffee", "pizza", "bakery",
        "burger", "burrito", "sushi", "ramen", "poke",
        "diner", "grill", "brewing",
        "starbucks", "tim horton", "mcdonald",
        "earls", "cactus", "white spot",
        "din tai fung", "rajio", "purebread", "loafe",
        "antico", "hot pie", "beach shack",
        "cloudburst", "tofitian", "jj bean", "dose coffee",
        "rain or shine", "regard coffee", "avik",
        "guilt & company", "sing sing",
        "steve's poke", "mucho burrito",
        "rosemary rocksalt", "shed restaurant",
        "angry otter",
        "fsm 0",
        "liquor",
    ],
    "Other Necessary": [
        "parking", "paybyphone",
        "doctor", "dr.", "clinic", "pharmacy", "physio", "health",
        "dentist", "optom",
        "car repair", "mechanic", "minit-tune", "brake",
        "ubc bookstore", "university of british",
        "evocarshare", "uber", "transit", "compass",
    ],
}

SKIP_KEYWORDS = [
    "payment - thank you", "paiement - merci",
    "payment received - thank you",
]

CUSTOM_FILE = "custom_categories.json"


# ── Persistent custom keyword storage ────────────────────────────

def load_custom(directory):
    path = os.path.join(directory, CUSTOM_FILE)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}


def save_custom(directory, data):
    with open(os.path.join(directory, CUSTOM_FILE), "w") as f:
        json.dump(data, f, indent=2)


# ── Classification logic ─────────────────────────────────────────

def auto_classify(description, customs):
    d = description.lower()
    for keyword, category in customs.items():
        if keyword in d:
            return category
    for category, keywords in AUTO_RULES.items():
        for kw in keywords:
            if kw in d:
                return category
    return None


def prompt_category(description, amount, customs, directory):
    print(f"\n  ? Cannot auto-classify:")
    print(f'    "{description}"  (${abs(amount):,.2f})')
    n = len(CATEGORIES)
    for i, cat in enumerate(CATEGORIES, 1):
        print(f"      {i}. {cat}")
    print(f"      {n + 1}. Skip this transaction")

    valid = {str(i) for i in range(1, n + 1)}
    while True:
        choice = input(f"    Choice [1-{n + 1}]: ").strip()
        if choice == str(n + 1):
            return None
        if choice in valid:
            cat = CATEGORIES[int(choice) - 1]
            kw = input("    Save a keyword for future auto-match? (or Enter to skip): ").strip()
            if kw:
                customs[kw.lower()] = cat
                save_custom(directory, customs)
                print(f"    Saved: '{kw}' -> {cat}")
            return cat
        print(f"    Invalid, try 1-{n + 1}.")


# ── RBC CSV parsing ──────────────────────────────────────────────

def parse_rbc_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            desc = row.get("Description 1", "").strip()
            cad = row.get("CAD$", "").strip()
            date_str = row.get("Transaction Date", "").strip()
            if not cad or not date_str:
                continue
            try:
                amount = float(cad)
                date = datetime.strptime(date_str, "%m/%d/%Y")
            except ValueError:
                continue
            rows.append({
                "date": date, "description": desc,
                "amount": amount, "source": "RBC",
            })
    return rows


# ── Amex XLS parsing ─────────────────────────────────────────────

def parse_amex_amount(text):
    """Parse '$19.35' or '-$2,832.00' into a float."""
    s = str(text).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_amex_date(text):
    """Parse '17 Mar. 2026' into a datetime."""
    s = str(text).strip().replace(".", "")
    if not s:
        return None
    try:
        return datetime.strptime(s, "%d %b %Y")
    except ValueError:
        return None


def parse_amex_xls(path):
    if xlrd is None:
        print(f"  Skipping {path} — xlrd not installed. Run: pip install xlrd")
        return []

    wb = xlrd.open_workbook(path)
    rows = []

    for sheet in wb.sheets():
        header_row = None
        for r in range(sheet.nrows):
            if sheet.cell_value(r, 0) == "Date":
                header_row = r
                break
        if header_row is None:
            continue

        for r in range(header_row + 1, sheet.nrows):
            date_val = str(sheet.cell_value(r, 0)).strip()
            desc_val = str(sheet.cell_value(r, 2)).strip()
            amount_val = str(sheet.cell_value(r, 3)).strip()

            # Amex payment rows put the amount in the Description column
            # and the description in Additional Information (col 9) or
            # Merchant Address (col 8)
            if not amount_val and desc_val.lstrip("-").startswith("$"):
                amount = parse_amex_amount(desc_val)
                desc_val = str(sheet.cell_value(r, 8)).strip()
                if not desc_val:
                    desc_val = str(sheet.cell_value(r, 9)).strip()
            else:
                amount = parse_amex_amount(amount_val)

            date = parse_amex_date(date_val)
            if date is None or amount is None or not desc_val:
                continue

            # Amex amounts are positive for charges, negative for payments/credits.
            # Normalize to RBC convention: negative = expense, positive = credit.
            rows.append({
                "date": date, "description": desc_val,
                "amount": -amount, "source": "Amex",
            })

    return rows


# ── Load all statements from a folder or single file ─────────────

def load_transactions(path):
    if os.path.isfile(path):
        if path.lower().endswith(".csv"):
            return parse_rbc_csv(path)
        elif path.lower().endswith(".xls"):
            return parse_amex_xls(path)
        else:
            print(f"Unsupported file type: {path}")
            return []

    if not os.path.isdir(path):
        print(f"Path not found: {path}")
        sys.exit(1)

    all_txns = []
    files = sorted(os.listdir(path))
    for fname in files:
        fpath = os.path.join(path, fname)
        if not os.path.isfile(fpath):
            continue
        if fname.lower().endswith(".csv"):
            loaded = parse_rbc_csv(fpath)
            print(f"  Loaded {len(loaded)} rows from {fname}  (RBC)")
            all_txns.extend(loaded)
        elif fname.lower().endswith(".xls"):
            loaded = parse_amex_xls(fpath)
            print(f"  Loaded {len(loaded)} rows from {fname}  (Amex)")
            all_txns.extend(loaded)

    return all_txns


def available_months(txns):
    return sorted({(t["date"].year, t["date"].month) for t in txns})


# ── Savings trajectory ───────────────────────────────────────────

def calc_savings(tfsa_contribution):
    today = datetime.now()
    months_left = max(
        (SAVINGS_DEADLINE.year - today.year) * 12
        + SAVINGS_DEADLINE.month - today.month,
        1,
    )
    gap = SAVINGS_GOAL - SAVINGS_BALANCE
    needed_monthly = gap / months_left
    projected = SAVINGS_BALANCE + tfsa_contribution * months_left
    return {
        "months_left": months_left,
        "needed_monthly": round(needed_monthly, 2),
        "projected": round(projected, 2),
        "on_track": projected >= SAVINGS_GOAL,
    }


# ── Console summary ──────────────────────────────────────────────

def print_summary(label, income, tfsa, totals, savings):
    total_exp = sum(totals.values())
    essential = (totals.get("Fixed Costs", 0)
                + totals.get("Work Travel (Island)", 0)
                + totals.get("Other Necessary", 0))
    available = income["total"] - essential - tfsa
    food_fun = totals.get("Food", 0) + totals.get("Discretionary/Fun", 0)
    remaining = available - food_fun

    w = 58
    print(f"\n{'=' * w}")
    print(f"   BUDGET SUMMARY  —  {label}")
    print(f"{'=' * w}")

    print(f"\n   INCOME")
    print(f"     Stipend               ${income['stipend']:>10,.2f}")
    print(f"     Ambulance             ${income['ambulance']:>10,.2f}")
    print(f"     Total                 ${income['total']:>10,.2f}")

    print(f"\n   SAVINGS  (20% of income)")
    print(f"     TFSA Contribution     ${tfsa:>10,.2f}")
    print(f"     Current Balance       ${SAVINGS_BALANCE:>10,.2f}  /  ${SAVINGS_GOAL:,.0f}")
    print(f"     Months Left           {savings['months_left']:>10}")
    print(f"     Needed / Month        ${savings['needed_monthly']:>10,.2f}")
    status = "On Track" if savings["on_track"] else "Behind"
    print(f"     Status                {status:>10}")

    print(f"\n   EXPENSES")
    for cat in CATEGORIES:
        print(f"     {cat:<26} ${totals.get(cat, 0):>10,.2f}")
    print(f"     {'─' * 40}")
    print(f"     {'Total':<26} ${total_exp:>10,.2f}")

    print(f"\n   LEFTOVER")
    print(f"     Available Fun Money   ${available:>10,.2f}")
    print(f"       (Income - Fixed - Travel - Other Nec. - TFSA)")
    print(f"     Spent (Food + Fun)    ${food_fun:>10,.2f}")
    print(f"     Remaining             ${remaining:>10,.2f}")
    if remaining < 0:
        print(f"     !! Over budget this month")
    print(f"\n{'=' * w}")

    return {"available": round(available, 2), "remaining": round(remaining, 2)}


# ── Excel output ──────────────────────────────────────────────────

def build_excel(path, label, income, tfsa, totals, savings, txns, leftover):
    wb = Workbook()
    cur = '#,##0.00'
    bold = Font(bold=True)
    section = Font(bold=True, size=11, color="2F5496")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    r = 1
    ws.cell(r, 1, f"Monthly Budget — {label}").font = Font(bold=True, size=14)

    r = 3
    ws.cell(r, 1, "INCOME").font = section
    r += 1
    ws.cell(r, 1, "Stipend"); ws.cell(r, 2, income["stipend"]).number_format = cur
    r += 1
    ws.cell(r, 1, "Ambulance Pay"); ws.cell(r, 2, income["ambulance"]).number_format = cur
    r += 1
    ws.cell(r, 1, "Total Income").font = bold
    c = ws.cell(r, 2, income["total"]); c.number_format = cur; c.font = bold

    r += 2
    ws.cell(r, 1, "SAVINGS  (20% of income)").font = section
    r += 1
    ws.cell(r, 1, "TFSA Contribution"); ws.cell(r, 2, tfsa).number_format = cur
    r += 1
    ws.cell(r, 1, "Current TFSA Balance"); ws.cell(r, 2, SAVINGS_BALANCE).number_format = cur
    r += 1
    ws.cell(r, 1, "Goal"); ws.cell(r, 2, f"${SAVINGS_GOAL:,.0f} by {SAVINGS_DEADLINE.strftime('%b %Y')}")
    r += 1
    ws.cell(r, 1, "Months Remaining"); ws.cell(r, 2, savings["months_left"])
    r += 1
    ws.cell(r, 1, "Needed per Month"); ws.cell(r, 2, savings["needed_monthly"]).number_format = cur
    r += 1
    ws.cell(r, 1, "Status"); ws.cell(r, 2, "On Track" if savings["on_track"] else "Behind")

    r += 2
    ws.cell(r, 1, "EXPENSES BY CATEGORY").font = section
    r += 1
    for cat in CATEGORIES:
        ws.cell(r, 1, cat); ws.cell(r, 2, totals.get(cat, 0)).number_format = cur
        r += 1
    total_exp = sum(totals.values())
    ws.cell(r, 1, "Total Spent").font = bold
    c = ws.cell(r, 2, total_exp); c.number_format = cur; c.font = bold

    r += 2
    ws.cell(r, 1, "BUDGET OVERVIEW").font = section
    r += 1
    ws.cell(r, 1, "Available Fun Money"); ws.cell(r, 2, leftover["available"]).number_format = cur
    r += 1
    ws.cell(r, 1, "  (Income - Fixed - Travel - Other Nec. - TFSA)")
    r += 1
    ws.cell(r, 1, "Spent on Food"); ws.cell(r, 2, totals.get("Food", 0)).number_format = cur
    r += 1
    ws.cell(r, 1, "Spent on Discretionary/Fun"); ws.cell(r, 2, totals.get("Discretionary/Fun", 0)).number_format = cur
    r += 1
    ws.cell(r, 1, "Remaining").font = bold
    c = ws.cell(r, 2, leftover["remaining"]); c.number_format = cur; c.font = bold
    if leftover["remaining"] < 0:
        c.font = Font(bold=True, color="FF0000")

    # ── Transactions sheet ──
    ws2 = wb.create_sheet("Transactions")
    ws2.column_dimensions["A"].width = 13
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 24
    ws2.column_dimensions["E"].width = 10

    headers = ["Date", "Description", "Amount", "Category", "Source"]
    for col, header in enumerate(headers, 1):
        c = ws2.cell(1, col, header)
        c.fill = hdr_fill
        c.font = hdr_font

    for i, t in enumerate(txns, 2):
        ws2.cell(i, 1, t["date"].strftime("%Y-%m-%d"))
        ws2.cell(i, 2, t["description"])
        ws2.cell(i, 3, round(-t["amount"], 2)).number_format = cur
        ws2.cell(i, 4, t["category"])
        ws2.cell(i, 5, t.get("source", ""))

    wb.save(path)


# ── Main ──────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("\nUsage:")
        print("  python budget_tracker.py <folder-of-statements>")
        print("  python budget_tracker.py <single-file.csv>")
        print("  python budget_tracker.py <single-file.xls>\n")
        print("Example:")
        print("  python budget_tracker.py credit_statements/\n")
        sys.exit(1)

    target = sys.argv[1]

    base_dir = os.path.dirname(os.path.abspath(__file__))
    customs = load_custom(base_dir)

    # ── Parse ──
    print("\nLoading statements...")
    txns = load_transactions(target)
    if not txns:
        print("No transactions found.")
        sys.exit(1)

    # ── Pick month ──
    months = available_months(txns)
    print(f"\n  {len(txns)} total transactions across {len(months)} month(s):\n")
    for i, (y, m) in enumerate(months, 1):
        n = sum(1 for t in txns if t["date"].year == y and t["date"].month == m)
        rbc_n = sum(1 for t in txns if t["date"].year == y and t["date"].month == m and t["source"] == "RBC")
        amex_n = sum(1 for t in txns if t["date"].year == y and t["date"].month == m and t["source"] == "Amex")
        parts = []
        if rbc_n:
            parts.append(f"{rbc_n} RBC")
        if amex_n:
            parts.append(f"{amex_n} Amex")
        detail = ", ".join(parts)
        print(f"  {i}. {datetime(y, m, 1).strftime('%B %Y')}  ({n} transactions: {detail})")

    if len(months) == 1:
        year, month = months[0]
    else:
        while True:
            try:
                idx = int(input("\nSelect month (number): ").strip()) - 1
                if 0 <= idx < len(months):
                    year, month = months[idx]
                    break
            except ValueError:
                pass
            print("Invalid choice.")

    label = datetime(year, month, 1).strftime("%B %Y")
    subset = [t for t in txns if t["date"].year == year and t["date"].month == month]

    # ── Filter out credit-card payments ──
    expenses = [
        t for t in subset
        if not any(kw in t["description"].lower() for kw in SKIP_KEYWORDS)
    ]

    # ── Ambulance income ──
    print(f"\n  Fixed stipend: ${STIPEND:,.2f}")
    while True:
        raw = input("  Ambulance pay this month (Enter for $0): $").strip()
        if raw == "":
            ambulance = 0.0
            break
        try:
            ambulance = float(raw.replace(",", ""))
            break
        except ValueError:
            print("  Invalid number.")

    income = {"stipend": STIPEND, "ambulance": ambulance, "total": STIPEND + ambulance}
    tfsa = round(income["total"] * SAVINGS_RATE, 2)

    # ── Categorize ──
    print(f"\n  Categorizing {len(expenses)} transactions...\n")
    categorized = []
    totals = {cat: 0.0 for cat in CATEGORIES}

    for t in expenses:
        if t["amount"] > 0 and "payment" in t["description"].lower():
            continue

        cat = auto_classify(t["description"], customs)

        if cat:
            tag = "  +"
        else:
            cat = prompt_category(t["description"], t["amount"], customs, base_dir)
            tag = "  *"

        if cat is None:
            print(f"    Skipped: {t['description']}")
            continue

        src = f"[{t['source']}]"
        print(f"{tag} {src:<7} {t['description'][:44]:<46} -> {cat}")
        t["category"] = cat
        categorized.append(t)
        totals[cat] -= t["amount"]

    # ── Savings trajectory ──
    savings = calc_savings(tfsa)

    # ── Print & write ──
    leftover = print_summary(label, income, tfsa, totals, savings)

    xlsx_path = os.path.join(base_dir, f"budget_{year}_{month:02d}.xlsx")
    build_excel(xlsx_path, label, income, tfsa, totals, savings, categorized, leftover)
    print(f"\n  Spreadsheet saved:  {xlsx_path}")
    print(f"  Upload to Google Drive -> Open with Google Sheets\n")


if __name__ == "__main__":
    main()
