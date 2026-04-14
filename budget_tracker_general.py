#!/usr/bin/env python3
"""
General-Purpose Monthly Budget Tracker
Reads RBC Visa (.csv) and Amex (.xls) credit card statements,
categorizes expenses via keyword matching, and generates a monthly
budget summary with Excel output for Google Sheets.

All financial details (income, savings goal, etc.) are entered
interactively and saved to a config file for reuse.
"""

import csv
import json
import os
import re
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Missing dependency. Run:  pip install openpyxl")
    sys.exit(1)

try:
    import xlrd
except ImportError:
    xlrd = None


# -- Default categories --------------------------------------------

DEFAULT_CATEGORIES = [
    "Transportation",
    "Fixed Costs",
    "Food",
    "Other Necessary",
    "Discretionary/Fun",
]

# ?? Auto-classification keywords (case-insensitive substring match) ??

AUTO_RULES = {
    "Transportation": [
        "bc ferries", "bcf -", "bcf-",
        "chevron", "chv",
        "shell c0",
        "petro-canada", "petro canada",
        "squamish valley gas",
        "uber", "lyft", "transit", "compass",
        "esso", "gas bar", "sunoco",
    ],
    "Fixed Costs": [
        "rent", "utilit", "wifi", "internet",
        "hydro", "bc hydro",
        "telus", "shaw", "rogers", "fido", "freedom mobile",
        "insurance", "bcaa",
        "interest",
        "mortgage",
    ],
    "Food": [
        "no frills", "superstore", "save-on", "safeway",
        "iga ", "walmart", "costco",
        "stong", "choices", "sungiven", "persia foods",
        "nesters", "grocery", "supermarket", "co-op",
        "cafe", "coffee", "pizza", "bakery",
        "burger", "burrito", "sushi", "ramen", "poke",
        "diner", "grill", "brewing",
        "starbucks", "tim horton", "mcdonald",
        "restaurant", "doordash", "skip the dishes", "ubereats",
        "liquor",
    ],
    "Other Necessary": [
        "parking", "paybyphone",
        "doctor", "dr.", "clinic", "pharmacy", "physio", "health",
        "dentist", "optom",
        "car repair", "mechanic", "brake",
        "bookstore",
    ],
}

SKIP_KEYWORDS = [
    "payment - thank you", "paiement - merci",
    "payment received - thank you",
]

CONFIG_FILE = "budget_config.json"
CUSTOM_FILE = "custom_categories.json"


# ?? Config management ?????????????????????????????????????????????

def prompt_float(prompt_text, default=None):
    while True:
        suffix = f" (Enter for ${default:,.2f})" if default is not None else ""
        raw = input(f"{prompt_text}{suffix}: $").strip().replace(",", "")
        if raw == "" and default is not None:
            return default
        try:
            return float(raw)
        except ValueError:
            print("  Please enter a valid number.")


def prompt_percentage(prompt_text, default=None):
    while True:
        suffix = f" (Enter for {default:.0%})" if default is not None else ""
        raw = input(f"{prompt_text}{suffix}: ").strip().rstrip("%")
        if raw == "" and default is not None:
            return default
        try:
            val = float(raw)
            if val > 1:
                val /= 100
            if 0 <= val <= 1:
                return val
            print("  Enter a value between 0 and 100.")
        except ValueError:
            print("  Please enter a valid number.")


def prompt_date(prompt_text, default=None):
    while True:
        suffix = f" (Enter for {default})" if default else ""
        raw = input(f"{prompt_text} [YYYY-MM]{suffix}: ").strip()
        if raw == "" and default:
            parts = default.split("-")
            return datetime(int(parts[0]), int(parts[1]), 1)
        try:
            parts = raw.split("-")
            return datetime(int(parts[0]), int(parts[1]), 1)
        except (ValueError, IndexError):
            print("  Use YYYY-MM format, e.g. 2027-09")


def load_config(directory):
    path = os.path.join(directory, CONFIG_FILE)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return None


def save_config(directory, cfg):
    with open(os.path.join(directory, CONFIG_FILE), "w") as f:
        json.dump(cfg, f, indent=2)


def run_setup(directory, existing=None):
    print("\n?? Budget Setup ??????????????????????????????????????")
    if existing:
        print("  (Press Enter to keep current values)\n")

    d = existing or {}

    monthly_income = prompt_float(
        "  Monthly income",
        default=d.get("monthly_income"),
    )
    savings_rate = prompt_percentage(
        "  Savings rate (e.g. 20 for 20%)",
        default=d.get("savings_rate"),
    )
    savings_goal = prompt_float(
        "  Savings goal",
        default=d.get("savings_goal"),
    )
    savings_balance = prompt_float(
        "  Current savings balance",
        default=d.get("savings_balance"),
    )
    savings_deadline = prompt_date(
        "  Savings deadline",
        default=d.get("savings_deadline"),
    )

    cfg = {
        "monthly_income": monthly_income,
        "savings_rate": savings_rate,
        "savings_goal": savings_goal,
        "savings_balance": savings_balance,
        "savings_deadline": savings_deadline.strftime("%Y-%m"),
    }
    save_config(directory, cfg)
    print("\n  Settings saved to budget_config.json\n")
    return cfg


# ?? Persistent custom keyword storage ????????????????????????????

def load_custom(directory):
    path = os.path.join(directory, CUSTOM_FILE)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}


def save_custom(directory, data):
    with open(os.path.join(directory, CUSTOM_FILE), "w") as f:
        json.dump(data, f, indent=2)


# ?? Classification logic ?????????????????????????????????????????

def auto_classify(description, customs):
    d = description.lower()
    for keyword, category in customs.items():
        if keyword in d:
            return category
    for category, keywords in AUTO_RULES.items():
        for kw in keywords:
            if kw in d:
                return category
    return None


def prompt_category(description, amount, customs, directory):
    categories = DEFAULT_CATEGORIES
    print(f"\n  ? Cannot auto-classify:")
    print(f'    "{description}"  (${abs(amount):,.2f})')
    n = len(categories)
    for i, cat in enumerate(categories, 1):
        print(f"      {i}. {cat}")
    print(f"      {n + 1}. Skip this transaction")

    valid = {str(i) for i in range(1, n + 1)}
    while True:
        choice = input(f"    Choice [1-{n + 1}]: ").strip()
        if choice == str(n + 1):
            return None
        if choice in valid:
            cat = categories[int(choice) - 1]
            kw = input("    Save a keyword for future auto-match? (or Enter to skip): ").strip()
            if kw:
                customs[kw.lower()] = cat
                save_custom(directory, customs)
                print(f"    Saved: '{kw}' -> {cat}")
            return cat
        print(f"    Invalid, try 1-{n + 1}.")


# ?? RBC CSV parsing ??????????????????????????????????????????????

def parse_rbc_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            desc = row.get("Description 1", "").strip()
            cad = row.get("CAD$", "").strip()
            date_str = row.get("Transaction Date", "").strip()
            if not cad or not date_str:
                continue
            try:
                amount = float(cad)
                date = datetime.strptime(date_str, "%m/%d/%Y")
            except ValueError:
                continue
            rows.append({
                "date": date, "description": desc,
                "amount": amount, "source": "RBC",
            })
    return rows


# ?? Amex XLS parsing ?????????????????????????????????????????????

def parse_amex_amount(text):
    s = str(text).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_amex_date(text):
    s = str(text).strip().replace(".", "")
    if not s:
        return None
    try:
        return datetime.strptime(s, "%d %b %Y")
    except ValueError:
        return None


def parse_amex_xls(path):
    if xlrd is None:
        print(f"  Skipping {path} — xlrd not installed. Run: pip install xlrd")
        return []

    wb = xlrd.open_workbook(path)
    rows = []

    for sheet in wb.sheets():
        header_row = None
        for r in range(sheet.nrows):
            if sheet.cell_value(r, 0) == "Date":
                header_row = r
                break
        if header_row is None:
            continue

        for r in range(header_row + 1, sheet.nrows):
            date_val = str(sheet.cell_value(r, 0)).strip()
            desc_val = str(sheet.cell_value(r, 2)).strip()
            amount_val = str(sheet.cell_value(r, 3)).strip()

            if not amount_val and desc_val.lstrip("-").startswith("$"):
                amount = parse_amex_amount(desc_val)
                desc_val = str(sheet.cell_value(r, 8)).strip()
                if not desc_val:
                    desc_val = str(sheet.cell_value(r, 9)).strip()
            else:
                amount = parse_amex_amount(amount_val)

            date = parse_amex_date(date_val)
            if date is None or amount is None or not desc_val:
                continue

            rows.append({
                "date": date, "description": desc_val,
                "amount": -amount, "source": "Amex",
            })

    return rows


# ?? Load all statements from a folder or single file ?????????????

def load_transactions(path):
    if os.path.isfile(path):
        if path.lower().endswith(".csv"):
            return parse_rbc_csv(path)
        elif path.lower().endswith(".xls"):
            return parse_amex_xls(path)
        else:
            print(f"Unsupported file type: {path}")
            return []

    if not os.path.isdir(path):
        print(f"Path not found: {path}")
        sys.exit(1)

    all_txns = []
    files = sorted(os.listdir(path))
    for fname in files:
        fpath = os.path.join(path, fname)
        if not os.path.isfile(fpath):
            continue
        if fname.lower().endswith(".csv"):
            loaded = parse_rbc_csv(fpath)
            print(f"  Loaded {len(loaded)} rows from {fname}  (RBC)")
            all_txns.extend(loaded)
        elif fname.lower().endswith(".xls"):
            loaded = parse_amex_xls(fpath)
            print(f"  Loaded {len(loaded)} rows from {fname}  (Amex)")
            all_txns.extend(loaded)

    return all_txns


def available_months(txns):
    return sorted({(t["date"].year, t["date"].month) for t in txns})


# ?? Savings trajectory ???????????????????????????????????????????

def calc_savings(cfg, monthly_contribution):
    today = datetime.now()
    parts = cfg["savings_deadline"].split("-")
    deadline = datetime(int(parts[0]), int(parts[1]), 1)
    balance = cfg["savings_balance"]
    goal = cfg["savings_goal"]

    months_left = max(
        (deadline.year - today.year) * 12
        + deadline.month - today.month,
        1,
    )
    gap = goal - balance
    needed_monthly = gap / months_left
    projected = balance + monthly_contribution * months_left
    return {
        "months_left": months_left,
        "needed_monthly": round(needed_monthly, 2),
        "projected": round(projected, 2),
        "on_track": projected >= goal,
        "goal": goal,
        "balance": balance,
        "deadline": deadline,
    }


# ?? Console summary ??????????????????????????????????????????????

def print_summary(label, income, savings_contribution, totals, savings):
    categories = DEFAULT_CATEGORIES
    total_exp = sum(totals.values())
    # Cash left after savings and all categorized card spending (same as income - savings - total_exp).
    remaining = income - savings_contribution - total_exp

    w = 58
    print(f"\n{'=' * w}")
    print(f"   BUDGET SUMMARY  —  {label}")
    print(f"{'=' * w}")

    print(f"\n   INCOME")
    print(f"     Monthly Income        ${income:>10,.2f}")

    rate_pct = f"{savings_contribution / income * 100:.0f}%" if income else "0%"
    print(f"\n   SAVINGS  ({rate_pct} of income)")
    print(f"     Monthly Contribution  ${savings_contribution:>10,.2f}")
    print(f"     Current Balance       ${savings['balance']:>10,.2f}  /  ${savings['goal']:,.0f}")
    print(f"     Months Left           {savings['months_left']:>10}")
    print(f"     Needed / Month        ${savings['needed_monthly']:>10,.2f}")
    status = "On Track" if savings["on_track"] else "Behind"
    print(f"     Status                {status:>10}")

    print(f"\n   EXPENSES")
    for cat in categories:
        print(f"     {cat:<26} ${totals.get(cat, 0):>10,.2f}")
    print(f"     {'-' * 40}")
    print(f"     {'Total':<26} ${total_exp:>10,.2f}")

    print(f"\n   LEFTOVER")
    print(f"     Remaining             ${remaining:>10,.2f}")
    if remaining < 0:
        print(f"     !! Over budget this month")
    print(f"\n{'=' * w}")

    return {"remaining": round(remaining, 2)}


# ?? Excel output ??????????????????????????????????????????????????

def build_excel(path, label, income, savings_contribution, totals, savings, txns, leftover):
    categories = DEFAULT_CATEGORIES
    wb = Workbook()
    cur = '#,##0.00'
    bold = Font(bold=True)
    section = Font(bold=True, size=11, color="2F5496")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    r = 1
    ws.cell(r, 1, f"Monthly Budget — {label}").font = Font(bold=True, size=14)

    r = 3
    ws.cell(r, 1, "INCOME").font = section
    r += 1
    ws.cell(r, 1, "Monthly Income").font = bold
    c = ws.cell(r, 2, income); c.number_format = cur; c.font = bold

    r += 2
    rate_pct = f"{savings_contribution / income * 100:.0f}%" if income else "0%"
    ws.cell(r, 1, f"SAVINGS  ({rate_pct} of income)").font = section
    r += 1
    ws.cell(r, 1, "Monthly Contribution"); ws.cell(r, 2, savings_contribution).number_format = cur
    r += 1
    ws.cell(r, 1, "Current Balance"); ws.cell(r, 2, savings["balance"]).number_format = cur
    r += 1
    ws.cell(r, 1, "Goal"); ws.cell(r, 2, f"${savings['goal']:,.0f} by {savings['deadline'].strftime('%b %Y')}")
    r += 1
    ws.cell(r, 1, "Months Remaining"); ws.cell(r, 2, savings["months_left"])
    r += 1
    ws.cell(r, 1, "Needed per Month"); ws.cell(r, 2, savings["needed_monthly"]).number_format = cur
    r += 1
    ws.cell(r, 1, "Status"); ws.cell(r, 2, "On Track" if savings["on_track"] else "Behind")

    r += 2
    ws.cell(r, 1, "EXPENSES BY CATEGORY").font = section
    r += 1
    for cat in categories:
        ws.cell(r, 1, cat); ws.cell(r, 2, totals.get(cat, 0)).number_format = cur
        r += 1
    total_exp = sum(totals.values())
    ws.cell(r, 1, "Total Spent").font = bold
    c = ws.cell(r, 2, total_exp); c.number_format = cur; c.font = bold

    r += 2
    ws.cell(r, 1, "LEFTOVER").font = section
    r += 1
    ws.cell(r, 1, "Remaining (income - savings - total expenses)").font = bold
    c = ws.cell(r, 2, leftover["remaining"]); c.number_format = cur; c.font = bold
    if leftover["remaining"] < 0:
        c.font = Font(bold=True, color="FF0000")

    ws2 = wb.create_sheet("Transactions")
    ws2.column_dimensions["A"].width = 13
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 24
    ws2.column_dimensions["E"].width = 10

    headers = ["Date", "Description", "Amount", "Category", "Source"]
    for col, header in enumerate(headers, 1):
        c = ws2.cell(1, col, header)
        c.fill = hdr_fill
        c.font = hdr_font

    for i, t in enumerate(txns, 2):
        ws2.cell(i, 1, t["date"].strftime("%Y-%m-%d"))
        ws2.cell(i, 2, t["description"])
        ws2.cell(i, 3, round(-t["amount"], 2)).number_format = cur
        ws2.cell(i, 4, t["category"])
        ws2.cell(i, 5, t.get("source", ""))

    wb.save(path)


# ?? Main ??????????????????????????????????????????????????????????

def main():
    if len(sys.argv) < 2:
        print("\nUsage:")
        print("  python budget_tracker_general.py <folder-of-statements>")
        print("  python budget_tracker_general.py <single-file.csv>")
        print("  python budget_tracker_general.py <single-file.xls>")
        print("\nFlags:")
        print("  --setup    Reconfigure income, savings, and goals\n")
        print("Example:")
        print("  python budget_tracker_general.py credit_statements/\n")
        sys.exit(1)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    reconfigure = "--setup" in sys.argv
    target = [a for a in sys.argv[1:] if a != "--setup"][0]

    cfg = load_config(base_dir)

    if cfg is None or reconfigure:
        cfg = run_setup(base_dir, existing=cfg)
    else:
        print(f"\n  Using saved settings (income: ${cfg['monthly_income']:,.2f}, "
              f"savings rate: {cfg['savings_rate']:.0%})")
        print(f"  Run with --setup to change these.\n")

    customs = load_custom(base_dir)

    # ?? Parse ??
    print("Loading statements...")
    txns = load_transactions(target)
    if not txns:
        print("No transactions found.")
        sys.exit(1)

    # ?? Pick month ??
    months = available_months(txns)
    print(f"\n  {len(txns)} total transactions across {len(months)} month(s):\n")
    for i, (y, m) in enumerate(months, 1):
        n = sum(1 for t in txns if t["date"].year == y and t["date"].month == m)
        rbc_n = sum(1 for t in txns if t["date"].year == y and t["date"].month == m and t["source"] == "RBC")
        amex_n = sum(1 for t in txns if t["date"].year == y and t["date"].month == m and t["source"] == "Amex")
        parts = []
        if rbc_n:
            parts.append(f"{rbc_n} RBC")
        if amex_n:
            parts.append(f"{amex_n} Amex")
        detail = ", ".join(parts)
        print(f"  {i}. {datetime(y, m, 1).strftime('%B %Y')}  ({n} transactions: {detail})")

    if len(months) == 1:
        year, month = months[0]
    else:
        while True:
            try:
                idx = int(input("\nSelect month (number): ").strip()) - 1
                if 0 <= idx < len(months):
                    year, month = months[idx]
                    break
            except ValueError:
                pass
            print("Invalid choice.")

    label = datetime(year, month, 1).strftime("%B %Y")
    subset = [t for t in txns if t["date"].year == year and t["date"].month == month]

    # ?? Filter out credit-card payments ??
    expenses = [
        t for t in subset
        if not any(kw in t["description"].lower() for kw in SKIP_KEYWORDS)
    ]

    income = cfg["monthly_income"]
    savings_contribution = round(income * cfg["savings_rate"], 2)

    # ?? Categorize ??
    print(f"\n  Categorizing {len(expenses)} transactions...\n")
    categorized = []
    totals = {cat: 0.0 for cat in DEFAULT_CATEGORIES}

    for t in expenses:
        if t["amount"] > 0 and "payment" in t["description"].lower():
            continue

        cat = auto_classify(t["description"], customs)

        if cat:
            tag = "  +"
        else:
            cat = prompt_category(t["description"], t["amount"], customs, base_dir)
            tag = "  *"

        if cat is None:
            print(f"    Skipped: {t['description']}")
            continue

        src = f"[{t['source']}]"
        print(f"{tag} {src:<7} {t['description'][:44]:<46} -> {cat}")
        t["category"] = cat
        categorized.append(t)
        totals[cat] -= t["amount"]

    # ?? Savings trajectory ??
    savings = calc_savings(cfg, savings_contribution)

    # ?? Print & write ??
    leftover = print_summary(label, income, savings_contribution, totals, savings)

    xlsx_path = os.path.join(base_dir, f"budget_{year}_{month:02d}.xlsx")
    build_excel(xlsx_path, label, income, savings_contribution, totals, savings, categorized, leftover)
    print(f"\n  Spreadsheet saved:  {xlsx_path}")
    print(f"  Upload to Google Drive -> Open with Google Sheets\n")


if __name__ == "__main__":
    main()
